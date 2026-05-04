#!/usr/bin/env python3
"""Build final consolidated audit report (all 3 passes)."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

p1 = json.load(open("/sessions/trusting-festive-ride/audit_pass1_issues.json"))
p2 = json.load(open("/sessions/trusting-festive-ride/audit_pass2_findings.json"))
p3 = json.load(open("/sessions/trusting-festive-ride/audit_pass3_findings.json"))

# Tag passes & filter meta files
META = {"chatbot-spec.md","system-prompt-v2.md"}
for i in p1: i["pass"] = "P1_structurel"
for i in p2: i["pass"] = "P2_factuel"
for i in p3: i["pass"] = "P3_editorial"

# Filter P1 meta noise
p1 = [i for i in p1 if i["page"] not in META]
# P2 normalize: P2 already has zone/kind/detail, ensure has no zone missing
for i in p2:
    if "zone" not in i: i["zone"] = ""
# P3 normalize
for i in p3:
    if "zone" not in i: i["zone"] = ""

all_issues = p1 + p2 + p3

# Priority computation
# HIGH = 3, MEDIUM = 2, LOW = 1
sev_score = {"HIGH":3,"MEDIUM":2,"LOW":1}
for i in all_issues:
    i["prio_score"] = sev_score.get(i["severity"],0)

# ─────────────────────────────────────
wb = Workbook()

header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
fill_h = PatternFill("solid", start_color="1F3864")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
sev_color = {"HIGH":"F4B084","MEDIUM":"FFE699","LOW":"C6EFCE"}
pass_color = {"P1_structurel":"D9E1F2","P2_factuel":"FCE4D6","P3_editorial":"E2EFDA"}

# ═══════════ Sheet 1 — Synthèse ═══════════
ws = wb.active
ws.title = "Synthèse"

ws["A1"] = "Audit Wiki Vancelian — Diagnostic complet"
ws["A1"].font = Font(bold=True, size=18, name="Arial", color="1F3864")
ws.merge_cells("A1:F1")

ws["A2"] = "Date : 18 avril 2026 — 3 passes (structurel · factuel · éditorial) — 236 fiches"
ws["A2"].font = Font(italic=True, size=11, name="Arial")
ws.merge_cells("A2:F2")

# Global KPIs
ws["A4"] = "Indicateurs globaux"
ws["A4"].font = Font(bold=True, size=13, name="Arial")
ws.merge_cells("A4:F4")

kpis = [
    ("Total anomalies détectées (3 passes)", len(all_issues)),
    ("  HIGH (bloquantes pour le bot)", sum(1 for x in all_issues if x["severity"]=="HIGH")),
    ("  MEDIUM (dégradantes)", sum(1 for x in all_issues if x["severity"]=="MEDIUM")),
    ("  LOW (cosmétiques)", sum(1 for x in all_issues if x["severity"]=="LOW")),
    ("", ""),
    ("Pages touchées par au moins 1 anomalie", len({x["page"] for x in all_issues})),
    ("Pages saines (0 anomalie)", 236 - len({x["page"] for x in all_issues})),
]
r = 5
for label, val in kpis:
    c1 = ws.cell(row=r, column=1, value=label); c1.font = Font(name="Arial", bold=(not label.startswith("  ")))
    c2 = ws.cell(row=r, column=2, value=val); c2.font = Font(name="Arial")
    r += 1

# By pass
ws[f"A{r+1}"] = "Répartition par passe"
ws[f"A{r+1}"].font = Font(bold=True, size=13, name="Arial")
r += 2

cols = ["Passe","HIGH","MEDIUM","LOW","Total","Focus"]
for j,c in enumerate(cols, start=1):
    cell = ws.cell(row=r, column=j, value=c)
    cell.font = header_font; cell.fill = fill_h; cell.alignment = center

pass_labels = {
    "P1_structurel": ("Frontmatter, sources, orphelins, liens, stale"),
    "P2_factuel": ("Marques, entités, AKTIO, JV Cloud Mining, Dubai, typos"),
    "P3_editorial": ("Short answer autonome, narration vs bullets, sections"),
}
r += 1
for pkey in ["P1_structurel","P2_factuel","P3_editorial"]:
    items = [x for x in all_issues if x["pass"]==pkey]
    hi = sum(1 for x in items if x["severity"]=="HIGH")
    me = sum(1 for x in items if x["severity"]=="MEDIUM")
    lo = sum(1 for x in items if x["severity"]=="LOW")
    ws.cell(row=r, column=1, value=pkey).fill = PatternFill("solid", start_color=pass_color[pkey])
    ws.cell(row=r, column=1).font = Font(name="Arial", bold=True)
    ws.cell(row=r, column=2, value=hi).font = Font(name="Arial")
    ws.cell(row=r, column=3, value=me).font = Font(name="Arial")
    ws.cell(row=r, column=4, value=lo).font = Font(name="Arial")
    ws.cell(row=r, column=5, value=hi+me+lo).font = Font(name="Arial", bold=True)
    ws.cell(row=r, column=6, value=pass_labels[pkey]).font = Font(name="Arial")
    r += 1

# Top 10 pages avec le plus d'anomalies
r += 2
ws.cell(row=r, column=1, value="Top 10 pages à prioriser (plus grand nombre d'anomalies × sévérité)").font = Font(bold=True, size=13, name="Arial")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
cols2 = ["Rang","Page","Score prio","# HIGH","# MEDIUM","# LOW"]
for j,c in enumerate(cols2, start=1):
    cell = ws.cell(row=r, column=j, value=c)
    cell.font = header_font; cell.fill = fill_h; cell.alignment = center

from collections import defaultdict
by_page = defaultdict(lambda: {"score":0,"H":0,"M":0,"L":0})
for i in all_issues:
    by_page[i["page"]]["score"] += i["prio_score"]
    if i["severity"]=="HIGH": by_page[i["page"]]["H"] += 1
    elif i["severity"]=="MEDIUM": by_page[i["page"]]["M"] += 1
    else: by_page[i["page"]]["L"] += 1

top10 = sorted(by_page.items(), key=lambda x:-x[1]["score"])[:10]
r += 1
for rank, (page, d) in enumerate(top10, start=1):
    ws.cell(row=r, column=1, value=rank).font = Font(bold=True, name="Arial")
    ws.cell(row=r, column=2, value=page).font = Font(name="Arial")
    ws.cell(row=r, column=3, value=d["score"]).font = Font(name="Arial", bold=True)
    ws.cell(row=r, column=4, value=d["H"]).font = Font(name="Arial")
    ws.cell(row=r, column=5, value=d["M"]).font = Font(name="Arial")
    ws.cell(row=r, column=6, value=d["L"]).font = Font(name="Arial")
    r += 1

# Col widths
for col, w in [("A",60),("B",55),("C",14),("D",10),("E",12),("F",55)]:
    ws.column_dimensions[col].width = w

# ═══════════ Sheet 2 — Plan d'action priorisé ═══════════
ws2 = wb.create_sheet("Plan d'action")
ws2["A1"] = "Plan de correction — par ordre de priorité"
ws2["A1"].font = Font(bold=True, size=16, name="Arial", color="1F3864")
ws2.merge_cells("A1:F1")

cols3 = ["#","Priorité","Titre","Ce qu'on fait","Pages concernées","Effort"]
for j,c in enumerate(cols3, start=1):
    cell = ws2.cell(row=3, column=j, value=c)
    cell.font = header_font; cell.fill = fill_h; cell.alignment = center

plan = [
    # (num, prio, titre, action, pages, effort)
    (1, "HIGH", "Typo entité légale 'Automata Pat Ltd'",
     "Corriger en 'Automata Pay Ltd' dans la page MiCA roadmap. Erreur grave car page à audience réglementaire.",
     "faq/legal-compliance/vancelian-mica-roadmap.md",
     "5 min"),
    (2, "HIGH", "Routage bot cassé sur 'can I withdraw AKTIO'",
     "Soit retirer la variante AKTIO des questions: de la page crypto-withdraw et la garder uniquement sur la page AKTIO dédiée, soit ajouter une phrase explicite dans la page crypto-withdraw : « AKTIO n'est pas withdrawable pour les clients EEA — voir la page AKTIO pour le cas ICO non-EEA. »",
     "faq/crypto/which-cryptoassets-can-i-withdraw-from-the-vancelian-applica.md",
     "15 min"),
    (3, "HIGH", "Dubai Villa : absence de 'Solaria' dans 2 fiches principales",
     "Ajouter une phrase d'ancrage : « L'emprunteur du prêt BTC est la société Solaria (SPV du promoteur). Le projet est localisé dans le quartier Al Barari à Dubaï. » Permet au bot de répondre correctement aux questions 'qui est Solaria', 'où est le projet'.",
     "faq/exclusive-offers/how-does-the-dubai-villa-al-barari-exclusive-offer-work.md ; faq/exclusive-offers/what-is-the-exclusive-offer-dubai-villa-al-barari.md",
     "20 min"),
    (4, "HIGH", "Ethiopia mining : incohérence Hearst entre 2 fiches du même produit",
     "Aligner la fiche 'how-does-the-exclusive-offer-eco-friendly-mining-in-ethiopia' avec 'what-is-the-eco-friendly-bitcoin-mining-in-ethiopia' (qui mentionne Hearst). Soit ajouter le partenaire Hearst dans les deux, soit confirmer qu'Ethiopia n'est PAS sous la JV Hearst (et alors retirer la mention de l'autre).",
     "faq/exclusive-offers/how-does-the-exclusive-offer-eco-friendly-mining-in-ethiopia.md",
     "30 min (à décider)"),
    (5, "HIGH", "17 citations sources: cassées (nommage Annexes + convention Top 30 + Dubai RTF)",
     "Normaliser : (a) mettre le vrai nom '_Automata France' dans les citations d'Annexes ; (b) pointer vers 'Top 30 questions bot Vancelian.docx' et utiliser un champ séparé pour le numéro de question (ex: note: 'Q23') ; (c) vérifier que 'Fiche_Jason_Dubai_Villa_EU.rtf' existe ou corriger la source.",
     "17 fiches (voir onglet Détail)",
     "1 h"),
    (6, "HIGH", "Short answer non-autonome (P3)",
     "Réécrire les short answers qui référencent 'voir ci-dessus/ci-dessous/liste' — elles doivent se tenir seules car le bot peut ne citer QUE cette section.",
     "Voir onglet Détail (P3, kind=short_answer_not_autonomous, si présent)",
     "selon volume"),
    (7, "MEDIUM", "Cloud Mining : absence de Hearst sur pages opérationnelles clés",
     "Ajouter un paragraphe d'ancrage JV sur les pages flow/CGUPM/early-exit : « Vancelian LTD est la JV entre Automata Group UK et Hearst Solution FZCO. Hearst fournit l'infrastructure de mining. » Permet au bot de répondre correctement aux questions sur la structure.",
     "faq/exclusive-offers/how-cloud-mining-flow-works.md ; cloud-mining-cgupm-investor-obligations.md ; cloud-mining-early-exit-and-transfers.md ; migration-to-the-new-cloud-mining-program.md",
     "45 min"),
    (8, "MEDIUM", "Fiches produit/risque dominées par bullets (18 pages, certaines >80%)",
     "Réécrire en narration privée banking (6 étapes) pour les pages les plus critiques : cloud-mining-risks-overview (95% bullets), cloud-mining-can-i-lose-my-capital (84%), the-heights-bali-project-reference (86%), aktio-utility-and-benefits (82%). Les autres peuvent rester en mode 'procédure' si c'est l'intention.",
     "18 pages (voir onglet Détail)",
     "2-3 h pour les 4 prioritaires"),
    (9, "MEDIUM", "Brand : mentions 'Akt.io' à contextualiser explicitement historique",
     "Sur 3 pages (what-is-the-story-of-vancelian, aktio-geographic-restrictions, aktio-utility-and-benefits), s'assurer que chaque mention 'Akt.io' est précédée d'un marqueur temporel explicite (« historiquement », « avant la refonte en Vancelian Q4 2024 »). Évite que le bot prenne Akt.io pour le nom actuel.",
     "3 pages",
     "20 min"),
    (10, "MEDIUM", "5 fiches sans section ## Details",
     "Ajouter la section ## Details manquante, même courte, pour respecter le schéma CLAUDE.md. Cela aide le bot à structurer sa réponse en 2 niveaux (short answer + détails).",
     "unauthorized-payment-on-my-card ; the-heights-bali-project-reference ; how-to-raise-my-deposit-limits ; how-to-open-a-vancelian-account ; mica-comprehensive-reference",
     "30 min"),
    (11, "MEDIUM", "1 lien related: cassé inter-catégorie",
     "Corriger le chemin relatif : dans faq/crypto/in-terms-of-valuation-potential-and-token-price-variation.md, remplacer 'where-is-aktio-listed.md' par '../aktio/where-is-aktio-listed.md'.",
     "faq/crypto/in-terms-of-valuation-potential-and-token-price-variation.md",
     "2 min"),
    (12, "LOW", "9 pages avec 11-16 variantes dans questions: (au-delà de 8 cible)",
     "Décision à prendre : soit tolérer car plus de variantes = meilleur retrieval Karpathy, soit normaliser à 8 pour discipline. Recommandation : tolérer pour les pages complexes (MiCA, Cloud Mining) et normaliser ailleurs.",
     "9 pages",
     "15 min si normalisation"),
    (13, "LOW", "5 short answers trop longues (>120 mots) sur offres exclusives",
     "Ces pages complexes peuvent légitimement dépasser 90 mots. Vérifier que l'essentiel tient en 3 phrases maximum pour la partie que le bot va quoter. Au-delà, déplacer en Details.",
     "5 pages (how-does-cloud-mining-by-hearst, btc-lending, dubai-villa-al-barari, guarantees-al-barari, cloud-mining-who-reimburses-if-bankruptcy)",
     "30 min"),
    (14, "LOW", "Exclure chatbot-spec.md et system-prompt-v2.md du périmètre wiki",
     "Les déplacer dans un dossier /meta/ ou les lister explicitement comme exclusions dans CLAUDE.md pour éviter les faux positifs aux audits futurs.",
     "2 fichiers",
     "5 min"),
]

r = 4
for num, prio, titre, action, pages, eff in plan:
    c_num = ws2.cell(row=r, column=1, value=num); c_num.font = Font(bold=True, name="Arial"); c_num.alignment = center
    c_prio = ws2.cell(row=r, column=2, value=prio)
    c_prio.font = Font(bold=True, name="Arial"); c_prio.fill = PatternFill("solid", start_color=sev_color.get(prio,"FFFFFF"))
    c_prio.alignment = center
    for col, val in [(3, titre),(4, action),(5, pages),(6, eff)]:
        c = ws2.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial")
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[r].height = 75
    r += 1

for col, w in [("A",5),("B",12),("C",35),("D",60),("E",45),("F",14)]:
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A4"

# ═══════════ Sheet 3 — Détail complet ═══════════
ws3 = wb.create_sheet("Détail complet")
cols4 = ["Page","Passe","Sévérité","Zone/Type","Détail","Extrait"]
for j,c in enumerate(cols4, start=1):
    cell = ws3.cell(row=1, column=j, value=c)
    cell.font = header_font; cell.fill = fill_h; cell.alignment = center

# Sort by sev then page
sev_order = {"HIGH":0,"MEDIUM":1,"LOW":2}
sorted_issues = sorted(all_issues, key=lambda x:(sev_order.get(x["severity"],3), x["page"]))

r = 2
for i in sorted_issues:
    ws3.cell(row=r, column=1, value=i["page"]).font = Font(name="Arial")
    c_pass = ws3.cell(row=r, column=2, value=i["pass"])
    c_pass.fill = PatternFill("solid", start_color=pass_color.get(i["pass"],"FFFFFF"))
    c_pass.font = Font(name="Arial")
    c_sev = ws3.cell(row=r, column=3, value=i["severity"])
    c_sev.fill = PatternFill("solid", start_color=sev_color.get(i["severity"],"FFFFFF"))
    c_sev.font = Font(bold=True, name="Arial"); c_sev.alignment = center
    zone_type = i.get("zone","") + (" · " if i.get("zone") and i.get("kind") else "") + i.get("kind","")
    ws3.cell(row=r, column=4, value=zone_type).font = Font(name="Arial")
    c_detail = ws3.cell(row=r, column=5, value=i.get("detail",""))
    c_detail.font = Font(name="Arial"); c_detail.alignment = Alignment(wrap_text=True, vertical="top")
    c_ex = ws3.cell(row=r, column=6, value=i.get("excerpt",""))
    c_ex.font = Font(name="Arial", italic=True, size=9); c_ex.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

for col, w in [("A",55),("B",14),("C",12),("D",30),("E",70),("F",60)]:
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = "A2"

# ═══════════ Sheet 4 — Lecture (méthode) ═══════════
ws4 = wb.create_sheet("Méthode")
ws4["A1"] = "Méthode d'audit"
ws4["A1"].font = Font(bold=True, size=16, name="Arial", color="1F3864")
ws4.merge_cells("A1:B1")

method_rows = [
    ("Périmètre", "236 fiches wiki markdown (faq/, entities/, concepts/, policies/) + raw/ pour vérification."),
    ("", ""),
    ("Passe 1 — Structurel (automatique)", "Scan de frontmatter YAML, validation des champs obligatoires (title, slug, category, questions, sources…), détection des orphelins d'index, liens related: cassés, sources raw/ introuvables, pages stale >6 mois, catégorie vs dossier, variantes questions: <5 ou >10."),
    ("", ""),
    ("Passe 2 — Factuel (ciblée zones à risque)", "Vérification factuelle sur 7 zones identifiées dans la mémoire projet : marques obsolètes (Akt.io/RAYN/Automata Pay), attribution d'entité (Vancelian LTD/Automata Group/Hearst/ADGM), AKTIO transférabilité (EEA/non-EEA), Cloud Mining JV + art. 4.3 CGUPM, type d'offre (mining/lending/refinancement), mécanique BTC générique, Dubai Villa Al Barari (quartier) / Solaria (emprunteur)."),
    ("", ""),
    ("Passe 3 — Éditorial", "Vérification méthode rédactionnelle : short answer autonome (pas de renvoi à d'autres sections), ratio narration vs listes à puces sur fiches produit/risque, présence des sections schéma (## Short answer, ## Details, ## Sources)."),
    ("", ""),
    ("Sévérité", "HIGH = casse une réponse du bot ou contredit un fait connu. MEDIUM = dégrade la qualité de la réponse ou la traçabilité. LOW = cosmétique ou question de discipline."),
    ("", ""),
    ("Faux positifs levés pendant l'audit", "(a) Encodage NFD/NFC sur noms de fichiers accentués (macOS) — normalisé. (b) Résolution des liens related: relatifs ('../') — corrigée. (c) Une page de reinvestissement générique flaggée à tort sur 'real estate' — déclassée."),
    ("", ""),
    ("Hors périmètre", "chatbot-spec.md et system-prompt-v2.md sont des méta-documents, pas des fiches wiki : exclus. Cohérence des TAUX chiffrés non vérifiée (requiert croisement manuel avec les brochures raw)."),
]
r = 3
for k, v in method_rows:
    ws4.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial", color="1F3864")
    ws4.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    ws4.cell(row=r, column=2, value=v).font = Font(name="Arial")
    ws4.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    if v: ws4.row_dimensions[r].height = 60
    r += 1

ws4.column_dimensions["A"].width = 35
ws4.column_dimensions["B"].width = 100

# Save
import os
out = "/sessions/trusting-festive-ride/mnt/Vancelian Support (Chat WIKI LLM)/audit-wiki/Audit_Wiki_Vancelian_Diagnostic_Complet.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"Saved: {out}")
